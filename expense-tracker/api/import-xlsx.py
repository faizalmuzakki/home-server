#!/usr/bin/env python3
"""Backfill the 2024/2025 "Expense Tracker" spreadsheets into the local SQLite db.

Usage: python3 import-xlsx.py <file.xlsx> [more.xlsx ...] [--db PATH] [--dry-run]

Only the 24 uniform month sheets are handled; trip/project sheets are bespoke and
are skipped by name (reported at the end). Rows land with source='xlsx', so:
    sqlite3 data/expenses.db "delete from expenses where source='xlsx'"
undoes the whole import, and re-running is idempotent.
"""
import argparse, calendar, datetime, os, re, sqlite3
import openpyxl

MONTH_SHEETS = {  # sheet name -> calendar month it is named for
    'Jan': 1, 'Feb': 2, 'March': 3, 'Apr': 4, 'April': 4, 'May': 5, 'June': 6,
    'July': 7, 'Aug': 8, 'August': 8, 'Sept': 9, 'Oct': 10, 'Nov': 11,
    'Dec': 12, 'Des': 12,
}
ID_MONTHS = {'januari': 1, 'februari': 2, 'maret': 3, 'april': 4, 'mei': 5, 'juni': 6,
             'juli': 7, 'agustus': 8, 'september': 9, 'oktober': 10, 'november': 11,
             'desember': 12}

# Spreadsheet category -> existing local category. Anything unlisted is created verbatim.
CATEGORY_MAP = {
    'Health': 'Healthcare',
    'Skin & Body (Care)': 'Skincare Bodycare',
    'Housing': 'Housing ',          # the local row really does have a trailing space
    'deoksi': 'deoksi.co',
    'Zakat': 'Sedekah/Zakat',
    'Zakat - sedekah': 'Sedekah/Zakat',
    'Self Development': 'Self Improvement',
    'byMuss': 'Bymuss',
}
INCOME_MAP = {}  # income sources kept verbatim (Paycheck Icha/Faizal, Side Hustle, Else)
BLANK_CATEGORY = 'Other'

# A category with a null colour blanks the dashboard: it feeds category.color straight
# into a recharts <Cell fill>, which calls .includes() on it and unmounts the tree.
FALLBACK_COLORS = ['#6366F1', '#EC4899', '#0EA5E9', '#F59E0B', '#8B5CF6', '#D946EF', '#059669', '#14B8A6']


def parse_day(v):
    """Cell -> (day, explicit_month, explicit_year); the last two are usually None."""
    if v is None or isinstance(v, bool) or v == '':
        return None, None, None
    if isinstance(v, datetime.datetime):
        return v.day, v.month, v.year
    if isinstance(v, (int, float)):
        return int(v), None, None
    s = str(v).strip()
    if s.isdigit():
        return int(s), None, None
    m = re.match(r'^(\d{1,2})\s+([A-Za-z]+)', s)          # e.g. "25 Februari"
    if m and m.group(2).lower() in ID_MONTHS:
        return int(m.group(1)), ID_MONTHS[m.group(2).lower()], None
    return None, None, None


def parse_amount(v):
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r'[^0-9.-]', '', str(v))
    try:
        return float(t)
    except ValueError:
        return None


def sheet_month(ws, name, fallback_year):
    """(year, month) the sheet is named for, from its anchor date cell if present."""
    for row in ws.iter_rows(min_row=1, max_row=12):
        for c in row:
            if isinstance(c.value, datetime.datetime):
                return c.value.year, c.value.month
    return fallback_year, MONTH_SHEETS[name.strip()]


def prev_month(year, month):
    return (year - 1, 12) if month == 1 else (year, month - 1)


def safe_date(year, month, day):
    """Clamp the day to the month's length; a few rows say 31 in a 30-day month."""
    return datetime.date(year, month, min(day, calendar.monthrange(year, month)[1])).isoformat()


def read_block(rows, hdr_i, day_col, year, month):
    """Yield (date, category, detail, note, amount) for one Day..Amount block."""
    hdr = rows[hdr_i]
    amt_col = next(j for j, h in enumerate(hdr) if h == 'Amount' and j > day_col)
    cat_col, det_col = day_col + 1, day_col + 2
    note_col = det_col + 1 if det_col + 1 < amt_col else None

    raw = []
    for r in rows[hdr_i + 1:]:
        amount = parse_amount(r[amt_col]) if amt_col < len(r) else None
        if amount is None:
            continue
        day, exp_m, exp_y = parse_day(r[day_col])
        if day is None:
            continue
        raw.append((day, exp_m, exp_y, r, amount))

    days = [x[0] for x in raw]
    # The sheets run on a payday cycle, not calendar months: each opens with the tail of
    # the previous month (the "2x" days at the top), rolls over to day 1, and closes with
    # this month's own 2x days. One rollover per sheet = first high -> low transition.
    roll = next((i for i in range(1, len(days)) if days[i - 1] >= 20 and days[i] <= 10), len(days))
    py, pm = prev_month(year, month)

    running_max = 0
    for i, (day, exp_m, exp_y, r, amount) in enumerate(raw):
        if i < roll:
            y, m = py, pm
        else:
            # A lone 29-31 landing right after the rollover is a late-entered row from the
            # previous month, not a jump to the end of this one. 13 rows across 24 sheets.
            straggler = day >= 25 and running_max < 15
            y, m = (py, pm) if straggler else (year, month)
            running_max = max(running_max, day)
        if exp_m:                       # "25 Februari" states the month outright; trust it
            m = exp_m
            y = exp_y or (year if exp_m <= month else py)
        cat = r[cat_col] if cat_col < len(r) else None
        det = r[det_col] if det_col < len(r) else None
        note = r[note_col] if note_col and note_col < len(r) else None
        yield (safe_date(y, m, day),
               str(cat).strip() if isinstance(cat, str) else '',
               str(det).strip() if isinstance(det, str) and det.strip() else None,
               str(note).strip() if isinstance(note, str) and note.strip() else None,
               amount)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('files', nargs='+')
    ap.add_argument('--db', default=os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                 'data/expenses.db'))
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    con = sqlite3.connect(args.db)
    cur = con.cursor()
    cat_ids, created, skipped_sheets = {}, [], []
    stats = {'expense': 0, 'income': 0, 'dupes': 0}
    per_sheet = []

    def category_id(name, kind):
        name = (name or '').strip() or (BLANK_CATEGORY if kind == 'expense' else 'Other Income')
        local = (CATEGORY_MAP if kind == 'expense' else INCOME_MAP).get(name, name)
        if local not in cat_ids:
            row = cur.execute('SELECT id FROM categories WHERE name = ?', (local,)).fetchone()
            if row:
                cat_ids[local] = row[0]
            else:
                cur.execute('INSERT INTO categories (name, type, icon, color) VALUES (?, ?, ?, ?)',
                            (local, kind, '\N{LABEL}', FALLBACK_COLORS[len(created) % len(FALLBACK_COLORS)]))
                cat_ids[local] = cur.lastrowid
                created.append(f'{local} ({kind})')
        return cat_ids[local]

    for path in args.files:
        year_hint = int(re.search(r'(20\d\d)', os.path.basename(path)).group(1))
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for name in wb.sheetnames:
            if name.strip() not in MONTH_SHEETS:
                skipped_sheets.append(f'{name}')
                continue
            ws = wb[name]
            rows = [[c.value for c in r] for r in ws.iter_rows()]
            hdr_i = next((i for i, r in enumerate(rows)
                          if 'Day' in [x for x in r if isinstance(x, str)]), None)
            if hdr_i is None:
                skipped_sheets.append(f'{name} (no Day header)')
                continue
            year, month = sheet_month(ws, name, year_hint)
            day_cols = [j for j, h in enumerate(rows[hdr_i]) if h == 'Day']
            n = {'expense': 0, 'income': 0}
            for k, day_col in enumerate(day_cols[:2]):
                kind = 'expense' if k == 0 else 'income'
                for date, cat, detail, note, amount in read_block(rows, hdr_i, day_col, year, month):
                    if cur.execute(
                            "SELECT 1 FROM expenses WHERE source='xlsx' AND date=? AND amount=?"
                            " AND description IS ? AND type=?",
                            (date, amount, detail, kind)).fetchone():
                        stats['dupes'] += 1
                        continue
                    cur.execute(
                        'INSERT INTO expenses (amount, description, category_id, date, type,'
                        ' source, raw_text) VALUES (?, ?, ?, ?, ?, ?, ?)',
                        (amount, detail, category_id(cat, kind), date, kind, 'xlsx', note))
                    stats[kind] += 1
                    n[kind] += 1
            per_sheet.append((os.path.basename(path)[:22], name, year, month, n['expense'], n['income']))
        wb.close()

    con.rollback() if args.dry_run else con.commit()

    for f, s, y, m, e, i in per_sheet:
        print(f'  {f:24} {s!r:10} -> {y}-{m:02d}  expense={e:4} income={i:4}')
    print(f'\n{stats}   {"(DRY RUN, rolled back)" if args.dry_run else ""}')
    print('new categories:', ', '.join(created) or 'none')
    print(f'skipped {len(skipped_sheets)} non-month sheets:', ', '.join(skipped_sheets))
    con.close()


if __name__ == '__main__':
    main()
